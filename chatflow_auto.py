from dotenv import load_dotenv
import os
from openai import OpenAI
from crews.input2cypher_crew import Input2CypherCrew
from crews.execute_crew import ExecuteCrew
from crews.generate_crew import GenerateCrew
from openpyxl import Workbook, load_workbook
import pandas as pd


# 读取环境变量
load_dotenv()
client = OpenAI(
    api_key=os.getenv("DEEPSEEK_API_KEY"),
    base_url="https://api.deepseek.com"
)

class ChatFlow:

    def __init__(self):
        self.retry_count = 0  # 失败重试计数器
        self.chat_history = []  # 存储对话历史

    def process_inputs(self, input_text: str):
        """ 处理用户输入，执行查询并记录对话历史 """
        self.chat_history.append({"role": "user", "content": input_text})  # 记录用户输入

        # 直接执行查询流程，不再判断是否是问候语
        context_content, response_content = self.chatflow_run(input_text)
        print(f"Chatbot: {response_content}")
        
        self.chat_history.append({"role": "assistant", "content": response_content})  # 记录机器人输出
        return context_content, response_content # 确保返回 context 和 response


    def write_to_excel(self, file_path, sheet_name, data, row, column):
        """将系统输出结果写入Excel，参与后续评估"""

        # 如果文件存在，加载已有的文件；否则新建一个
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
        else:
            wb = Workbook()
            ws = wb.active
            if sheet_name:
                ws.title = sheet_name

        # 写入数据
        ws.cell(row=row, column=column, value=data)
        wb.save(file_path)


    def chatflow_run(self, input_text: str):
        """ 执行完整的查询流程，并支持重试 """
        crew1 = Input2CypherCrew(input_text)
        query = crew1.run()

        crew2 = ExecuteCrew(f"生成的Cypher查询如下：{query}，请借助工具向Neo4j数据库执行查询")
        query_answer = crew2.run()

        if query_answer.startswith("查询失败"):
            self.retry_count += 1
            if self.retry_count < 3:
                print(f"Chatbot: 查询失败，正在重新尝试...(第 {self.retry_count} 次)")
                return self.chatflow_run(input_text)
            else:
                self.retry_count = 0
                return "查询失败，请调整您的问题或稍后再试。", "经过多次尝试，查询仍然失败，请调整您的问题或稍后再试。"

        # 查询成功，生成自然语言结果
        self.retry_count = 0
        crew3 = GenerateCrew(f"输入问题为：{input_text}"
                             f"你将接受到一个Json格式的输出结果：\n{query_answer}，请根据用户输入和系统输出，优化输出为符合人类语言习惯的中文结果。")
        output = crew3.run()
        return query_answer, output


if __name__ == "__main__":
    chat_flow = ChatFlow()  # 机器人实例，保持对话上下文
    df = pd.read_excel('评价表.xlsx')
    # 循环读取input，然后输出response、context
    for i in range(df.shape[0]):
        user_input = df.loc[i, 'input']
        context, response = chat_flow.process_inputs(user_input)
        # 写入系统输出
        chat_flow.write_to_excel('评价表.xlsx', 'Sheet1', response, i + 2, 4)
        # 写入上下文
        chat_flow.write_to_excel('评价表.xlsx', 'Sheet1', context, i + 2, 5)