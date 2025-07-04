import re, os
from openpyxl import Workbook, load_workbook


def process_string(input_str):
    
    data = re.sub(r'[\[\]{}]', '', input_str)
    output_data = re.sub(r"[\"']", r'\"', data)
    return output_data

# 注意:openpylx使用的row,cloumn与真实Excel行列号一一对应
def write_to_excel(file_path, sheet_name, data, row, column):
    # 如果文件存在，加载已有的文件；否则新建一个
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        if sheet_name:
            ws.title = sheet_name

    # 写入数据
    ws.cell(row=row, column=column, value=data)
    wb.save(file_path)
     

if __name__ == '__main__':
    input_str = "[{'b':{'identity':371035736,'labels':['Building'],'properties':{'gml:id':'bldg_ac8d4b0e-6f5b-4caf-b9c8-9af3a82685ac','type':'Building','region':'1丰岛区'},'elementId':'4:be6dcfeb-90a4-4ae6-b3aa-1ccc9172af42:371035736'},'height': '234.1', 'location': '東京都豊島区東池袋三丁目'}]"
    result = process_string(input_str)
    write_to_excel('evaluation_table.xlsx', 'Sheet1', result, 2, 5)